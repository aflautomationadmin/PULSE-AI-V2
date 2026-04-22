import openpyxl
from openpyxl.styles import (
    PatternFill, Font, Alignment, Border, Side, GradientFill
)
from openpyxl.utils import get_column_letter

wb = openpyxl.Workbook()

# ── colour palette ──────────────────────────────────────────────────────────
C_HEADER_BG   = "1F3864"   # dark navy
C_HEADER_FG   = "FFFFFF"
C_DONE_BG     = "E2EFDA"   # light green
C_PENDING_BG  = "FFF2CC"   # light yellow
C_INPROG_BG   = "DDEBF7"   # light blue
C_ALT_ROW     = "F5F5F5"   # subtle grey for alternating rows
C_SECTION_BG  = "D6E4F0"   # section header row
C_TITLE_BG    = "2E75B6"   # title bar

STATUS_COLORS = {
    "Done":        "70AD47",  # green
    "In Progress": "2E75B6",  # blue
    "Pending":     "ED7D31",  # orange
    "Cancelled":   "A6A6A6",  # grey
}

def hdr_fill(hex_code):
    return PatternFill("solid", fgColor=hex_code)

def thin_border():
    s = Side(style="thin", color="BFBFBF")
    return Border(left=s, right=s, top=s, bottom=s)

def status_badge(ws, row, col, status):
    cell = ws.cell(row=row, column=col, value=status)
    color = STATUS_COLORS.get(status, "A6A6A6")
    cell.fill = PatternFill("solid", fgColor=color)
    cell.font = Font(bold=True, color="FFFFFF", size=10)
    cell.alignment = Alignment(horizontal="center", vertical="center")
    cell.border = thin_border()

# ============================================================
# SHEET 1 – TASK TRACKER
# ============================================================
ws = wb.active
ws.title = "Task Tracker"

# ── Title bar ───────────────────────────────────────────────
ws.merge_cells("A1:H1")
t = ws["A1"]
t.value = "AI-DA-AGENTS  ·  Project Task Tracker"
t.fill = hdr_fill(C_TITLE_BG)
t.font = Font(bold=True, color="FFFFFF", size=16)
t.alignment = Alignment(horizontal="center", vertical="center")
ws.row_dimensions[1].height = 36

ws.merge_cells("A2:H2")
sub = ws["A2"]
sub.value = "Retail Sales Analytics Chatbot  |  Last updated: 06-Apr-2026"
sub.fill = hdr_fill("3E87C8")
sub.font = Font(color="FFFFFF", size=10, italic=True)
sub.alignment = Alignment(horizontal="center", vertical="center")
ws.row_dimensions[2].height = 18

# ── Column headers ───────────────────────────────────────────
headers = ["#", "Category", "Task Description", "File(s) Changed",
           "Status", "Priority", "Completed On", "Notes"]
ws.row_dimensions[3].height = 22
for col, h in enumerate(headers, start=1):
    c = ws.cell(row=3, column=col, value=h)
    c.fill = hdr_fill(C_HEADER_BG)
    c.font = Font(bold=True, color=C_HEADER_FG, size=11)
    c.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
    c.border = thin_border()

# ── Column widths ────────────────────────────────────────────
col_widths = [4, 22, 55, 48, 14, 10, 16, 48]
for i, w in enumerate(col_widths, start=1):
    ws.column_dimensions[get_column_letter(i)].width = w

# ── Task data ────────────────────────────────────────────────
# (num, category, description, files_changed, status, priority, completed_on, notes)
SECTIONS = {
    "Foundation": [],
    "Visualization": [],
    "Agent Enhancements": [],
    "Entity Resolution": [],
    "SQL Intelligence": [],
    "Bug Fixes": [],
}

tasks = [
    # Foundation
    ("Foundation", 1,  "Replicate AIbot-Agents project into AI-DA-AGENTS",
     "All src/ files",
     "Done", "High", "Mar-2026",
     "Full project copy; confirmed identical structure in both repos"),

    # Visualization
    ("Visualization", 2,  "Generate styled HTML table when user requests 'table'",
     "src/visualization.py, app.py",
     "Done", "High", "Mar-2026",
     "render_table_html() added; table label fixed in app.py"),
    ("Visualization", 3,  "Fix pie chart always rendering as bar chart",
     "src/visualization.py",
     "Done", "High", "Mar-2026",
     "Added explicit_pie token check before semantic heuristics"),
    ("Visualization", 4,  "Fix line chart when no date column present",
     "src/visualization.py",
     "Done", "Medium", "Mar-2026",
     "Falls back to label_index when date_index is None"),
    ("Visualization", 5,  "Fix bar chart explicit detection",
     "src/visualization.py",
     "Done", "Medium", "Mar-2026",
     "Added explicit_bar / explicit_chart token layers"),
    ("Visualization", 6,  "Replace keyword-driven auto-chart with data-driven rule",
     "src/visualization.py",
     "Done", "High", "Mar-2026",
     "Removed _COMPARE_KEYWORDS; chart only when row_count >= 2"),
    ("Visualization", 7,  "Prevent HTML table creation for single-row scalar results",
     "src/visualization.py",
     "Done", "Medium", "Mar-2026",
     "Guard: if len(result.rows) >= _MIN_ROWS_FOR_AUTO_CHART"),

    # Agent Enhancements
    ("Agent Enhancements", 8,  "Add Clarifier agent — ask follow-up for vague/open-ended questions",
     "src/agents/clarifier.py (new), src/orchestrator.py",
     "Done", "High", "Mar-2026",
     "Checks: time period (critical), dimension, metric. Reads history to avoid re-asking"),
    ("Agent Enhancements", 9,  "Add Empty Result Handler — smart follow-up when SQL returns 0 rows",
     "src/agents/empty_result_handler.py (new), src/orchestrator.py",
     "Done", "High", "Apr-2026",
     "LLM acknowledges no data, states likely reason, asks ONE corrective question"),
    ("Agent Enhancements", 10, "Enhance Empty Result Handler to cite resolved entity attributes",
     "src/agents/empty_result_handler.py, src/orchestrator.py",
     "Done", "Medium", "06-Apr-2026",
     "Passes self.last_entity_match; LLM says e.g. 'I searched SUBCLASS=JOGGERS but found nothing'"),

    # Entity Resolution
    ("Entity Resolution", 11, "Add STATE entity resolver (embedding + fuzzy)",
     "src/entity_search.py, src/orchestrator.py",
     "Done", "High", "Mar-2026",
     "StateEntityResolver; threshold from settings"),
    ("Entity Resolution", 12, "Add CITY entity resolver",
     "src/entity_search.py, src/orchestrator.py",
     "Done", "High", "Mar-2026",
     "CityEntityResolver; shares cache with state resolver"),
    ("Entity Resolution", 13, "Add STORE_NAME entity resolver",
     "src/entity_search.py, src/orchestrator.py",
     "Done", "High", "Mar-2026",
     "StoreNameEntityResolver"),
    ("Entity Resolution", 14, "Add CATEGORY entity resolver with value embeddings",
     "src/entity_search.py, src/orchestrator.py",
     "Done", "High", "Mar-2026",
     "CategoryEntityResolver; threshold capped at 0.80"),
    ("Entity Resolution", 15, "Add SUBCLASS entity resolver (lower threshold for misspellings)",
     "src/entity_search.py, src/orchestrator.py",
     "Done", "High", "Apr-2026",
     "SubclassEntityResolver; threshold 0.78; catches 'Jogers' → JOGGERS"),

    # SQL Intelligence
    ("SQL Intelligence", 16, "Add brand vs category/subclass disambiguation rule for 'polo'",
     "src/agents/sql_writer.py, business_context.json",
     "Done", "High", "Mar-2026",
     "'polo' alone → CATEGORY/SUBCLASS; 'USPA/US POLO' → BRAND"),
    ("SQL Intelligence", 17, "Add KPI definitions (ABV, ASP, ABS, ATV) to SQL writer",
     "src/agents/sql_writer.py",
     "Done", "High", "Mar-2026",
     "Full Unique Bills formula with INVOICETYPE CASE logic"),
    ("SQL Intelligence", 18, "Add growth KPI rule (MoM/YoY/WoW) to SQL writer",
     "src/agents/sql_writer.py",
     "Done", "Medium", "Mar-2026",
     "Returns current, previous, absolute change, % change"),
    ("SQL Intelligence", 19, "Add SQL debugger agent with auto-retry on DB error",
     "src/agents/sql_debugger.py, src/orchestrator.py",
     "Done", "High", "Mar-2026",
     "_execute_with_recovery(); retries = sql_debug_max_retries"),
    ("SQL Intelligence", 20, "Add semantic SQL cache (exact + cosine similarity)",
     "src/sql_cache.py, src/orchestrator.py",
     "Done", "High", "Mar-2026",
     "SQLite store; fingerprint on schema + business context; embedding lookup"),

    # Bug Fixes
    ("Bug Fixes", 21, "Fix 'polo' resolved as USPA brand instead of product category",
     "src/agents/sql_writer.py, business_context.json",
     "Done", "High", "Mar-2026",
     "Disambiguation rule + USPA note updated"),
    ("Bug Fixes", 22, "Fix chart label showing raw type instead of capitalised name",
     "app.py",
     "Done", "Low", "Mar-2026",
     "label = 'Table' if chart_type == 'table' else chart_type.capitalize()"),
    ("Bug Fixes", 23, "Fix single-row questions always creating an HTML table file",
     "src/visualization.py",
     "Done", "Medium", "Mar-2026",
     "Row count guard added to build_visual_output()"),
]

row = 4
current_section = None

for (section, num, desc, files, status, priority, completed, notes) in tasks:
    # Insert section header row when section changes
    if section != current_section:
        current_section = section
        ws.merge_cells(f"A{row}:H{row}")
        sc = ws.cell(row=row, column=1, value=f"  {section.upper()}")
        sc.fill = hdr_fill(C_SECTION_BG)
        sc.font = Font(bold=True, color="1F3864", size=11)
        sc.alignment = Alignment(vertical="center")
        sc.border = thin_border()
        ws.row_dimensions[row].height = 20
        row += 1

    # Alternating row fill (behind status badge cells)
    bg = C_DONE_BG if status == "Done" else (C_INPROG_BG if status == "In Progress" else C_PENDING_BG)

    data = [num, section, desc, files, status, priority, completed, notes]
    for col, val in enumerate(data, start=1):
        c = ws.cell(row=row, column=col, value=val)
        c.border = thin_border()
        c.alignment = Alignment(vertical="center", wrap_text=True,
                                horizontal="center" if col in (1, 5, 6, 7) else "left")
        c.font = Font(size=10)
        if col != 5:   # skip status cell — handled separately
            c.fill = PatternFill("solid", fgColor=bg if col > 1 else "FFFFFF")

    # Status badge
    status_badge(ws, row, 5, status)

    # Priority colour
    prio_cell = ws.cell(row=row, column=6)
    prio_colors = {"High": "C00000", "Medium": "ED7D31", "Low": "70AD47"}
    prio_cell.font = Font(bold=True, color=prio_colors.get(priority, "000000"), size=10)

    ws.row_dimensions[row].height = 42
    row += 1

# ── Freeze panes ─────────────────────────────────────────────
ws.freeze_panes = "A4"

# ── Auto-filter ──────────────────────────────────────────────
ws.auto_filter.ref = f"A3:H{row - 1}"


# ============================================================
# SHEET 2 – SUMMARY DASHBOARD
# ============================================================
ws2 = wb.create_sheet("Summary")

# Title
ws2.merge_cells("A1:E1")
t2 = ws2["A1"]
t2.value = "Project Summary Dashboard"
t2.fill = hdr_fill(C_TITLE_BG)
t2.font = Font(bold=True, color="FFFFFF", size=14)
t2.alignment = Alignment(horizontal="center", vertical="center")
ws2.row_dimensions[1].height = 30

# Count tasks by status
from collections import Counter
status_counts = Counter(t[4] for t in tasks)
category_counts = Counter(t[0] for t in tasks)
total = len(tasks)

# Status summary table
ws2.cell(row=3, column=1, value="STATUS SUMMARY").font = Font(bold=True, size=12, color="1F3864")
hdr_row = 4
for col, h in enumerate(["Status", "Count", "% of Total"], start=1):
    c = ws2.cell(row=hdr_row, column=col, value=h)
    c.fill = hdr_fill(C_HEADER_BG)
    c.font = Font(bold=True, color="FFFFFF")
    c.alignment = Alignment(horizontal="center")
    c.border = thin_border()

r = 5
for status, count in sorted(status_counts.items()):
    ws2.cell(row=r, column=1, value=status).border = thin_border()
    ws2.cell(row=r, column=2, value=count).border = thin_border()
    pct = ws2.cell(row=r, column=3, value=f"{count/total*100:.0f}%")
    pct.border = thin_border()
    color = STATUS_COLORS.get(status, "A6A6A6")
    for col in range(1, 4):
        c2 = ws2.cell(row=r, column=col)
        c2.fill = PatternFill("solid", fgColor=color)
        c2.font = Font(color="FFFFFF", bold=True)
        c2.alignment = Alignment(horizontal="center")
    r += 1

# Totals row
ws2.cell(row=r, column=1, value="TOTAL").font = Font(bold=True)
ws2.cell(row=r, column=2, value=total).font = Font(bold=True)
ws2.cell(row=r, column=3, value="100%").font = Font(bold=True)
for col in range(1, 4):
    ws2.cell(row=r, column=col).border = thin_border()

# Category breakdown table
ws2.cell(row=3, column=5, value="TASKS BY CATEGORY").font = Font(bold=True, size=12, color="1F3864")
for col, h in enumerate(["Category", "Tasks", "Done"], start=5):
    c = ws2.cell(row=4, column=col, value=h)
    c.fill = hdr_fill(C_HEADER_BG)
    c.font = Font(bold=True, color="FFFFFF")
    c.alignment = Alignment(horizontal="center")
    c.border = thin_border()

r2 = 5
done_by_cat = Counter(t[0] for t in tasks if t[4] == "Done")
for cat, cnt in sorted(category_counts.items()):
    done = done_by_cat.get(cat, 0)
    ws2.cell(row=r2, column=5, value=cat).border = thin_border()
    ws2.cell(row=r2, column=6, value=cnt).border = thin_border()
    d_cell = ws2.cell(row=r2, column=7, value=done)
    d_cell.border = thin_border()
    if done == cnt:
        d_cell.fill = PatternFill("solid", fgColor="70AD47")
        d_cell.font = Font(color="FFFFFF", bold=True)
    for col in (5, 6, 7):
        ws2.cell(row=r2, column=col).alignment = Alignment(horizontal="center")
    r2 += 1

# Column widths for sheet 2
for col, w in [(1, 18), (2, 10), (3, 12), (4, 4), (5, 24), (6, 10), (7, 10)]:
    ws2.column_dimensions[get_column_letter(col)].width = w


# ============================================================
# SHEET 3 – CHANGE LOG
# ============================================================
ws3 = wb.create_sheet("Change Log")
ws3.merge_cells("A1:F1")
tl = ws3["A1"]
tl.value = "Change Log"
tl.fill = hdr_fill(C_TITLE_BG)
tl.font = Font(bold=True, color="FFFFFF", size=14)
tl.alignment = Alignment(horizontal="center", vertical="center")
ws3.row_dimensions[1].height = 30

log_headers = ["Date", "Task #", "Change Description", "File(s)", "Author", "Status"]
for col, h in enumerate(log_headers, start=1):
    c = ws3.cell(row=2, column=col, value=h)
    c.fill = hdr_fill(C_HEADER_BG)
    c.font = Font(bold=True, color="FFFFFF")
    c.alignment = Alignment(horizontal="center")
    c.border = thin_border()

log_entries = [
    ("Mar-2026", 1,  "Initial project replication from AIbot-Agents",             "All src/",                              "Dev", "Done"),
    ("Mar-2026", 2,  "Added render_table_html(); fixed app.py chart label",        "visualization.py, app.py",              "Dev", "Done"),
    ("Mar-2026", 3,  "Fixed pie chart detection (explicit token check)",           "visualization.py",                      "Dev", "Done"),
    ("Mar-2026", 4,  "Fixed line chart fallback when no date column",              "visualization.py",                      "Dev", "Done"),
    ("Mar-2026", 5,  "Added explicit bar chart detection",                         "visualization.py",                      "Dev", "Done"),
    ("Mar-2026", 6,  "Replaced keyword-driven chart logic with row-count rule",    "visualization.py",                      "Dev", "Done"),
    ("Mar-2026", 7,  "Fixed single-row scalar results creating HTML table",        "visualization.py",                      "Dev", "Done"),
    ("Mar-2026", 8,  "Added ClarifierAgent for vague questions",                   "clarifier.py (new), orchestrator.py",   "Dev", "Done"),
    ("Mar-2026", 11, "Added StateEntityResolver",                                  "entity_search.py, orchestrator.py",     "Dev", "Done"),
    ("Mar-2026", 12, "Added CityEntityResolver",                                   "entity_search.py, orchestrator.py",     "Dev", "Done"),
    ("Mar-2026", 13, "Added StoreNameEntityResolver",                              "entity_search.py, orchestrator.py",     "Dev", "Done"),
    ("Mar-2026", 14, "Added CategoryEntityResolver (threshold 0.80)",              "entity_search.py, orchestrator.py",     "Dev", "Done"),
    ("Mar-2026", 16, "Added polo/USPA disambiguation rule",                        "sql_writer.py, business_context.json",  "Dev", "Done"),
    ("Mar-2026", 17, "Added KPI definitions to sql_writer instructions",           "sql_writer.py",                         "Dev", "Done"),
    ("Mar-2026", 18, "Added growth KPI (MoM/YoY/WoW) rule",                       "sql_writer.py",                         "Dev", "Done"),
    ("Mar-2026", 19, "Added SQL debugger agent + auto-retry",                      "sql_debugger.py, orchestrator.py",      "Dev", "Done"),
    ("Mar-2026", 20, "Added semantic SQL cache (exact + cosine)",                  "sql_cache.py, orchestrator.py",         "Dev", "Done"),
    ("Apr-2026",  9, "Added EmptyResultHandler agent (0-row follow-up)",           "empty_result_handler.py, orchestrator.py", "Dev", "Done"),
    ("Apr-2026", 15, "Added SubclassEntityResolver (threshold 0.78)",              "entity_search.py, orchestrator.py",     "Dev", "Done"),
    ("06-Apr-2026", 10, "Enhanced EmptyResultHandler with entity attribute context", "empty_result_handler.py, orchestrator.py", "Dev", "Done"),
]

for r, entry in enumerate(log_entries, start=3):
    for col, val in enumerate(entry, start=1):
        c = ws3.cell(row=r, column=col, value=val)
        c.border = thin_border()
        c.alignment = Alignment(vertical="center", wrap_text=True,
                                horizontal="center" if col in (1, 2, 5, 6) else "left")
        c.font = Font(size=10)
        if r % 2 == 0:
            c.fill = PatternFill("solid", fgColor=C_ALT_ROW)
    ws3.row_dimensions[r].height = 18

for col, w in [(1, 14), (2, 8), (3, 55), (4, 40), (5, 12), (6, 12)]:
    ws3.column_dimensions[get_column_letter(col)].width = w

ws3.auto_filter.ref = f"A2:F{2 + len(log_entries)}"

# ── Save ─────────────────────────────────────────────────────
out = r"C:\Users\7518549\WORK\AI-DA-AGENTS\AI_DA_AGENTS_Task_Tracker.xlsx"
wb.save(out)
print(f"Saved: {out}")
